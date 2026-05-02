#!/usr/bin/env python3
"""
Veracode Tenant-Wide Scan Health
Python port of https://github.com/veracode/scan_health (Go v2.47)
Extended for tenant-wide iteration, Excel/CSV/JSON export, trend analysis.

Requirements:
    pip install veracode-api-signing requests openpyxl
"""

from __future__ import annotations

import csv
import json
import re
import time
import logging
import argparse
import threading
import xml.etree.ElementTree as ET
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Callable

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger("scan_health")

# ==========================================================================
# Constants
# ==========================================================================
MAX_FILE_COUNT = 10000
MAX_MODULE_COUNT = 500
MAX_SELECTED_MODULE_COUNT = 100
MAX_FLAW_COUNT = 2500
MAX_TOTAL_MODULE_SIZE = 1_000_000_000
MAX_ANALYSIS_SIZE = 500_000_000
STALE_SCAN_DAYS = 30

REGIONS: dict[str, dict[str, str]] = {
    "commercial": {"base": "https://analysiscenter.veracode.com", "xml": "https://analysiscenter.veracode.com/api/5.0", "rest": "https://api.veracode.com/appsec/v1"},
    "eu":         {"base": "https://analysiscenter.veracode.eu",  "xml": "https://analysiscenter.veracode.eu/api/5.0",  "rest": "https://api.veracode.eu/appsec/v1"},
}

_NS = re.compile(r'\sxmlns="[^"]+"')
_DT_FMTS = ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S %Z", "%Y-%m-%d %H:%M:%S UTC")
_URL_RE = re.compile(r'https?://\S+')

# FancyList pattern lists
JUNK_FILE_PATTERNS = ["!LICENSE*",".*","*.asmx","*.config","*.cs","*.eot","*.gif","*.ico","*.jpeg","*.jpg","*.less","*.manifest","*.map","*.markdown","*.md","*.pdf","*.png","*.properties","*.scss","*.sh","*.svg","*.ttf","*.txt","*.woff","*.xml","AUTHORS","CHANGELOG","CONTRIBUTORS","Dockerfile","LICENSE","Makefile","README","Thumbs.db"]
THIRD_PARTY_PATTERNS = ["7z.dll","7-zip.dll","Google.*.dll","7za.exe","AutoMapper.dll","AutoMapper.*.dll","Azure.*.dll","BouncyCastle.*","Castle.Core.*","Castle.Windsor.*","componentspace.saml2.dll","Dapper.dll","Dapper.*.dll","devexpress.*","entityframework.*","Google.Protobuf.dll","gradle-wrapper.jar","GraphQL.*.dll","itextsharp.dll","log4net.dll","microsoft.*.dll","microsoft.*.pdb","!^_Microsoft.","!^_System.","!^_Azure.","newrelic.*.dll","newtonsoft.json.*","ninject.*.dll","org.eclipse.*.jar","Serilog.dll","syncfusion.*","system.*.dll","Telerik.*.dll","WebGrease.dll","phantomjs.exe","Moq.dll","ComponentSpace.SAML2.dll","^aspnet-codegenerator","sni.dll","AntiXssLibrary.dll","Antlr3.Runtime.dll","FluentValidation.dll"]
SENSITIVE_SECRET_PATTERNS = ["*.asc","*.crt","*.gpg","*.jks","*.key","*.p7b","*.p7s","*.pem","*.pfx","*.pgp","*.p12","*.tfvars","variable.tf",".htpasswd"]
SENSITIVE_BACKUP_PATTERNS = ["*.bac","*.back","*.backup","*.old","*.orig","*.bak"]
SENSITIVE_WORD_PATTERNS = ["*.docx","*.doc","*.docm","*.odt"]
SENSITIVE_SPREADSHEET_PATTERNS = ["*.xlsx","*.xls","*.xlsm","*.ods"]
SENSITIVE_JUPYTER_PATTERNS = ["*.ipynb"]
TEST_FILE_PATTERNS = ["nunit.framework.dll","Moq.dll","^.test.","!Test*","!*Test","!^Test.","!^Tests.","*.unittests.dll","*.unittest.dll","^mock","^unittest","^harness","*.feature","*.js.snap"]
UNWANTED_7Z = ["*.7z"]
UNWANTED_COFFEE = ["*.coffee"]
UNWANTED_SCRIPTS = ["*.sh","*.ps","*.ps1","*.bat"]
UNWANTED_INSTALLERS = ["setup.exe","*setup.exe","*.msi","installer.exe","*installer.exe","*.msix","*.appx","*.msixbundle",".appxbundle"]
UNWANTED_PYD = ["*.pyd"]
UNWANTED_PYC = ["*.pyc"]
UNWANTED_DEPLOY = ["*.deploy"]
UNWANTED_WIBU = ["WibuCmNET.dll"]
SRC_JAVA = ["*.java"]; SRC_CS = ["*.cs"]; SRC_SLN = ["*.sln"]; SRC_CSPROJ = ["*.csproj"]; SRC_C = ["*.c"]; SRC_CPP = ["*.cpp"]; SRC_SWIFT = ["*.swift"]
SCA_SUPPORTED = ["*.dll","*.exe","*.jar","*.apk","*.aab","*.war","*.ear","*.js","*.ts","*.php","*.lock","package-lock.json","npm-shrinkwrap.json","go.sum","vendor.json","*.deps.json","*.py"]
DOTNET_PRECOMPILE_PATTERNS = ["*.cshtml","*.ascx","*.aspx","*.asax"]
GO_WORKSPACE_PATTERNS = ["go.work","go.work.sum"]
EXCESS_MSFT_PATTERNS = ["csc.exe"]
LOOSE_CLASS_FILE = ["*.class"]; LOOSE_CLASS_MODULE = ["class files within*"]
REPO_CANARIES = ["fsmonitor-watchman.sample","FETCH_HEAD"]
GRADLE_WRAPPER = ["gradle-wrapper.jar"]; MINIFIED_JS = ["*.min.js"]

# Excel styling
_HF = PatternFill("solid", fgColor="1F4E79")
_HN = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_DF = Font(name="Arial", size=10)
_BF = Font(bold=True, name="Arial", size=10)
_CA = Alignment(horizontal="center", vertical="center")
_WA = Alignment(horizontal="left", vertical="top", wrap_text=True)
_TH = Side(style="thin", color="CCCCCC")
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
_CLR = {"Good": "C6EFCE", "Fair": "FFEB9C", "Poor": "FFC7CE"}
_AGE_CLR = {"<7d": "C6EFCE", "7-30d": "C6EFCE", "30-90d": "FFEB9C", "90d+": "FFC7CE", "N/A": "FFFFFF"}

# ==========================================================================
# Dataclasses
# ==========================================================================

@dataclass
class Issue:
    severity: str
    description: str
    check_num: int = 0
    check_name: str = ""
    check_num: int = 0
    check_name: str = ""

CHECK_CATEGORIES: dict[int, str] = {
    1: "Packaging", 2: "Module Selection", 3: "Flaw Analysis",
    4: "Fatal Errors", 5: "Fatal Errors", 6: "Packaging",
    7: "Packaging", 8: "Packaging", 9: "SCA",
    10: "Module Selection", 11: "Packaging", 12: "Module Quality",
    13: "Module Quality", 14: "Fatal Errors", 15: "Module Selection",
    16: "Security Risk", 17: "Security Risk", 18: "Packaging",
    19: "Packaging", 20: "Packaging", 21: "Packaging",
    22: "Packaging", 23: "Packaging", 24: "Module Selection",
    25: "Module Selection", 26: "Module Selection", 27: "Packaging",
    28: "Packaging", 29: "Module Selection", 30: "Scan Recency",
    31: "Packaging",
}

@dataclass
class FlawSummary:
    total: int = 0; fixed: int = 0; pol_aff: int = 0
    mitigated: int = 0; open_pol: int = 0; open_nopol: int = 0

@dataclass
class ScanResult:
    app_name: str = ""; bu: str = ""; policy: str = ""; app_id: int = 0
    sandbox: str = ""; build_id: str = ""; scan_name: str = ""
    is_latest: bool = True; scan_status: str = "No Scan"
    published: str = ""; days_since: object = "N/A"; duration: str = ""
    engine: str = ""; analysis_size_mb: float = 0.0
    files_uploaded: int = 0; total_modules: int = 0
    selected_modules: int = 0; fatal_errors: int = 0
    flaws: FlawSummary = field(default_factory=FlawSummary)
    health: str = "Good"; high_issues: int = 0; medium_issues: int = 0
    total_issues: int = 0; issues_text: str = "None"; recs_text: str = "None"
    review_url: str = ""; triage_url: str = ""
    selected_names: str = ""; sca_count: int = 0; age_bucket: str = "N/A"
    total_upload_mb: float = 0.0; health_trend: str = ""

    def to_row(self) -> dict:
        return {
            "App Name": self.app_name, "Business Unit": self.bu, "Policy": self.policy,
            "App ID": self.app_id, "Sandbox": self.sandbox,
            "Build ID": self.build_id, "Scan Name": self.scan_name,
            "Is Latest": self.is_latest, "Scan Status": self.scan_status,
            "Published": self.published, "Days Since Scan": self.days_since,
            "Duration": self.duration, "Engine": self.engine,
            "Analysis Size (MB)": self.analysis_size_mb,
            "Total Upload Size (MB)": self.total_upload_mb,
            "Files Uploaded": self.files_uploaded,
            "Total Modules": self.total_modules,
            "Selected Modules": self.selected_modules,
            "Selected Module Names": self.selected_names,
            "Fatal Errors": self.fatal_errors,
            "Total Flaws": self.flaws.total,
            "Open Affecting Policy": self.flaws.open_pol,
            "Mitigated": self.flaws.mitigated, "Fixed": self.flaws.fixed,
            "Policy Affecting": self.flaws.pol_aff,
            "SCA Components": self.sca_count,
            "Scan Age Bucket": self.age_bucket,
            "Health": self.health, "Health Trend": self.health_trend,
            "High Issues": self.high_issues, "Medium Issues": self.medium_issues,
            "Total Issues": self.total_issues,
            "Issues": self.issues_text, "Recommendations": self.recs_text,
            "Review Modules URL": self.review_url, "Triage Flaws URL": self.triage_url,
        }

@dataclass
class ModuleRow:
    app_name: str = ""; build_id: str = ""; name: str = ""
    status: str = ""; selected: bool = False; dependency: bool = False
    fatal: bool = False; third_party: bool = False; ignored: bool = False
    platform: str = ""; compiler: str = ""; size: str = ""
    issues: str = ""
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Build ID": self.build_id, "Module": self.name,
                "Status": self.status, "Selected": self.selected, "Dependency": self.dependency,
                "Fatal": self.fatal, "3rd Party": self.third_party, "Ignored": self.ignored,
                "Platform": self.platform, "Compiler": self.compiler, "Size": self.size,
                "Issues": self.issues}

@dataclass
class FileRow:
    app_name: str = ""; build_id: str = ""; name: str = ""
    status: str = ""; md5: str = ""; ignored: bool = False; third_party: bool = False
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Build ID": self.build_id, "File": self.name,
                "Status": self.status, "MD5": self.md5, "Ignored": self.ignored, "3rd Party": self.third_party}

@dataclass
class RecommendationRow:
    app_name: str = ""; build_id: str = ""; severity: str = ""
    recommendation: str = ""; doc_url: str = ""
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Build ID": self.build_id,
                "Severity": self.severity, "Recommendation": self.recommendation,
                "Doc URL": self.doc_url}

@dataclass
class TrendRow:
    app_name: str = ""; sandbox: str = ""
    prev_health: str = ""; curr_health: str = ""; change: str = ""
    prev_flaws: int = 0; curr_flaws: int = 0; flaw_delta: int = 0
    prev_open_pol: int = 0; curr_open_pol: int = 0; open_pol_delta: int = 0
    def to_row(self) -> dict:
        return {"App Name": self.app_name, "Sandbox": self.sandbox,
                "Previous Health": self.prev_health, "Current Health": self.curr_health,
                "Health Change": self.change,
                "Previous Total Flaws": self.prev_flaws, "Current Total Flaws": self.curr_flaws,
                "Flaw Delta": self.flaw_delta,
                "Previous Open Policy": self.prev_open_pol, "Current Open Policy": self.curr_open_pol,
                "Open Policy Delta": self.open_pol_delta}

@dataclass
class AppIssues:
    """Carries structured issue data per app for aggregation."""
    app_name: str = ""
    bu: str = ""
    policy: str = ""
    sandbox: str = ""
    health: str = ""
    issues: list[Issue] = field(default_factory=list)
    recs: list[str] = field(default_factory=list)

@dataclass
class AggIssue:
    """Structured issue record for accurate tenant-level aggregation."""
    app_name: str
    bu: str
    sandbox: str
    check_num: int
    check_name: str
    category: str
    severity: str
    description: str
    recommendation: str

# ==========================================================================
# Helpers
# ==========================================================================

def _si(v: object, d: int = 0) -> int:
    try: return int(v)
    except (ValueError, TypeError): return d

def _parse_dt(s: str) -> datetime | None:
    if not s: return None
    s = s.strip()
    for f in _DT_FMTS:
        try:
            dt = datetime.strptime(s, f)
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except ValueError: continue
    return None

def _days_since(s: str) -> int | None:
    dt = _parse_dt(s)
    return (datetime.now(timezone.utc) - dt).days if dt else None

def _dur(sub: str, pub: str) -> str:
    a, b = _parse_dt(sub), _parse_dt(pub)
    if a and b:
        s = int((b - a).total_seconds())
        if s >= 0:
            h, r = divmod(s, 3600); m, sec = divmod(r, 60)
            if h: return f"{h}h {m}m {sec}s"
            return f"{m}m {sec}s" if m else f"{sec}s"
    return ""

def _age_bucket(days: int | None) -> str:
    if days is None: return "N/A"
    if days < 7: return "<7d"
    if days <= 30: return "7-30d"
    if days <= 90: return "30-90d"
    return "90d+"

def _parse_module_size(size_str: str) -> int:
    """Parse prescan module size string like '5MB', '120KB', '2GB' to bytes."""
    s = size_str.strip().upper()
    for suffix, mult in [("GB", 1_000_000_000), ("MB", 1_000_000), ("KB", 1_000)]:
        if s.endswith(suffix):
            try: return int(s[:-len(suffix)]) * mult
            except ValueError: return 0
    return 0

def _extract_url(text: str) -> str:
    m = _URL_RE.search(text)
    return m.group(0).rstrip(".") if m else ""

def _is_java(n: str) -> bool: return n.lower().endswith((".jar",".war",".ear"))
def _is_dotnet(n: str) -> bool: return n.lower().endswith((".dll",".exe"))
def _is_js_module(n: str) -> bool:
    lo = n.lower()
    return lo.startswith("js files within") or lo.startswith("js files extracted from")
def _is_node_module(n: str) -> bool: return "_nodemodule_" in n.lower()
def _has_status(m: dict, s: str) -> bool: return s.lower() in m.get("status","").lower()

def _top5(items: list[str]) -> str:
    if len(items) <= 5: return ", ".join(f'"{i}"' for i in items)
    return ", ".join(f'"{i}"' for i in items[:5]) + f" and {len(items)-5} others"

# ==========================================================================
# FancyList DSL (audited against Go utils/fancy_list.go)
# ==========================================================================

def _fancy_match(filename: str, patterns: list[str]) -> bool:
    """Test a filename against a list of patterns using the FancyList DSL.

    Pattern syntax:
      - Default: case-insensitive equality match (e.g. "readme")
      - Single *: wildcard. "*.exe" matches suffix, "abc.*" matches prefix,
                  "abc.*.xyz" matches prefix+suffix. Only one * allowed.
      - ^ prefix: contains search (e.g. "^test" matches if "test" in filename)
      - ! prefix: force case-sensitive matching for the rest of the pattern.
                  Can combine with * or ^ (e.g. "!^_Microsoft." is case-sensitive
                  contains, "!Test*" is case-sensitive prefix).
    """
    fn = filename.strip()
    for pat in patterns:
        p = pat
        # ! means case-sensitive; otherwise fold to lowercase
        case_sensitive = p.count("!") == 1
        if case_sensitive:
            p = p.replace("!", "")
            f = fn
        else:
            f = fn.lower()
            p = p.lower()
        # ^ means contains
        if "^" in p and p.count("^") == 1:
            p = p.replace("^", "")
            if p in f:
                return True
            continue
        # * wildcard (exactly one allowed)
        if p.count("*") == 1:
            if p.startswith("*"):
                if f.endswith(p[1:]):
                    return True
            elif p.endswith("*"):
                if f.startswith(p[:-1]):
                    return True
            else:
                parts = p.split("*", 1)
                if f.startswith(parts[0]) and f.endswith(parts[1]) and len(f) >= len(parts[0]) + len(parts[1]):
                    return True
        elif f == p:
            return True
    return False

def _fancy_match_files(files: list[dict], patterns: list[str]) -> list[str]:
    found: list[str] = []
    for f in files:
        if f.get("is_ignored") or f.get("is_third_party"): continue
        if _fancy_match(f["name"], patterns) and f["name"] not in found:
            found.append(f["name"])
    return found

def _fancy_match_modules(modules: list[dict], patterns: list[str], selected_only: bool = False) -> list[str]:
    found: list[str] = []
    for m in modules:
        if selected_only and not m.get("is_selected"): continue
        if _fancy_match(m["name"], patterns) and m["name"] not in found:
            found.append(m["name"])
    return found

# ==========================================================================
# API Client
# ==========================================================================

class AuthError(Exception):
    """Raised on 401/403 to signal credential issues."""

class VeracodeClient:
    def __init__(self, region: str = "commercial", timeout: int = 120) -> None:
        self._cfg = REGIONS[region]
        self._timeout = timeout
        self._s = requests.Session()
        self._s.auth = RequestsAuthPluginVeracodeHMAC()
        self._s.headers["User-Agent"] = "veracode-scan-health-py/3.0"
        retry = Retry(total=3, backoff_factor=1.0,
                      status_forcelist=(429, 500, 502, 503, 504),
                      allowed_methods=("GET",))
        self._s.mount("https://", HTTPAdapter(max_retries=retry))

    def close(self) -> None: self._s.close()
    def __enter__(self) -> "VeracodeClient": return self
    def __exit__(self, *a: object) -> None: self.close()

    def _check_auth(self, resp: requests.Response) -> None:
        if resp.status_code in (401, 403):
            body = resp.text[:200] if resp.text else ""
            raise AuthError(
                f"HTTP {resp.status_code} from {resp.url}. "
                f"API credentials may be expired or lack required permissions. "
                f"Response: {body}")

    def _xml(self, ep: str, params: dict | None = None) -> ET.Element:
        r = self._s.get(f"{self._cfg['xml']}/{ep}", params=params, timeout=self._timeout)
        self._check_auth(r)
        r.raise_for_status()
        # Strip ALL xmlns attributes (count=0) for clean XPath
        return ET.fromstring(_NS.sub("", r.text, count=0))

    def _rest(self, path: str, params: dict | None = None) -> dict:
        r = self._s.get(f"{self._cfg['rest']}{path}", params=params, timeout=self._timeout)
        self._check_auth(r)
        r.raise_for_status()
        return r.json()

    @property
    def base(self) -> str: return self._cfg["base"]

    def get_apps(self) -> list[dict]:
        apps: list[dict] = []; page = 0
        while True:
            d = self._rest("/applications", {"page": page, "size": 500})
            emb = d.get("_embedded", {}).get("applications", [])
            if not emb: break
            for a in emb:
                p = a.get("profile", {}); pols = p.get("policies") or []
                apps.append({"app_id": a.get("guid",""), "legacy_id": a.get("id"),
                    "name": p.get("name",""),
                    "bu": (p.get("business_unit") or {}).get("name",""),
                    "policy": pols[0].get("name","") if pols else ""})
            page += 1
            if page >= (d.get("page") or {}).get("total_pages", 1): break
        return apps

    def get_builds(self, aid: int, sbx: str | None = None) -> list[dict]:
        p: dict = {"app_id": aid}
        if sbx: p["sandbox_id"] = sbx
        try: root = self._xml("getbuildlist.do", p)
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getbuildlist.do failed for %s: %s", aid, e); return []
        return [{"id": b.get("build_id"), "ver": b.get("version","")} for b in root.findall(".//build")]

    def get_sandboxes(self, aid: int) -> list[dict]:
        try: root = self._xml("getsandboxlist.do", {"app_id": aid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getsandboxlist.do failed: %s", e); return []
        return [{"id": s.get("sandbox_id",""), "name": s.get("sandbox_name","")} for s in root.findall(".//sandbox")]

    def get_build_info(self, aid: int, bid: str) -> dict | None:
        """Get build status to determine if scan is published."""
        try: root = self._xml("getbuildinfo.do", {"app_id": aid, "build_id": bid})
        except (requests.HTTPError, ET.ParseError): return None
        au = root.find(".//analysis_unit")
        if au is None: return None
        return {"status": au.get("status",""), "published": au.get("published_date","")}

    def get_detailed_report(self, bid: str) -> dict | None:
        try: root = self._xml("detailedreport.do", {"build_id": bid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("detailedreport.do failed for %s: %s", bid, e); return None
        sa = root.find(".//static-analysis")
        if sa is None: return None
        fl = FlawSummary()
        for f in root.findall(".//severity/category/cwe/staticflaws/flaw"):
            fl.total += 1
            apc = f.get("affects_policy_compliance","false") == "true"
            is_fixed = f.get("remediation_status","") == "Fixed"
            is_miti = f.get("mitigation_status","none") not in ("none","rejected")
            if apc: fl.pol_aff += 1
            if is_fixed: fl.fixed += 1
            elif is_miti: fl.mitigated += 1
            elif apc: fl.open_pol += 1
            else: fl.open_nopol += 1
        if fl.total == 0:
            fl.total = _si(root.get("total_flaws","0"))
            fl.open_pol = _si(root.get("flaws_not_mitigated","0"))
            fl.mitigated = fl.total - fl.open_pol
        dr_mods = [{"name": unescape(m.get("name","")), "compiler": unescape(m.get("compiler","")),
            "os": unescape(m.get("os","")), "arch": unescape(m.get("architecture",""))}
            for m in (sa.findall(".//module") or [])]
        sca_node = root.find(".//software_composition_analysis")
        sca_on = sca_node is not None and sca_node.get("sca_service_available","true").lower() != "false"
        sca_comps = [unescape(c.get("file_name","")) for c in root.findall(".//vulnerable_components/component")]
        return {
            "account_id": root.get("account_id",""), "app_id": root.get("app_id",""),
            "sandbox_id": _si(root.get("sandbox_id","0")), "sandbox_name": root.get("sandbox_name",""),
            "analysis_id": root.get("analysis_id",""), "sau_id": root.get("static_analysis_unit_id",""),
            "bu": unescape(root.get("business_unit","")), "app_name": unescape(root.get("app_name","")),
            "scan_name": unescape(sa.get("version","")), "engine": sa.get("engine_version",""),
            "submitted": sa.get("submitted_date",""), "published": sa.get("published_date",""),
            "analysis_size": _si(sa.get("analysis_size_bytes","0")),
            "is_latest": root.get("is_latest_build","true").lower() == "true",
            "flaws": fl, "dr_modules": dr_mods,
            "sca_on": sca_on, "sca_comps": sca_comps,
        }

    def get_files(self, aid: int, bid: str) -> list[dict]:
        try: root = self._xml("getfilelist.do", {"app_id": aid, "build_id": bid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getfilelist.do failed: %s", e); return []
        return [{"name": unescape(f.get("file_name","")), "status": unescape(f.get("file_status","")),
                 "md5": f.get("file_md5",""), "is_ignored": False, "is_third_party": False}
                for f in root.findall(".//file")]

    def get_prescan(self, aid: int, bid: str) -> list[dict]:
        try: root = self._xml("getprescanresults.do", {"app_id": aid, "build_id": bid})
        except (requests.HTTPError, ET.ParseError) as e:
            log.debug("getprescanresults.do failed: %s", e); return []
        mods: list[dict] = []
        for m in root.findall(".//module"):
            issues: list[str] = []
            for iss in m.findall(".//issue"):
                d = unescape(iss.get("details",""))
                if d and d not in issues: issues.append(d)
            st = unescape(m.get("status",""))
            if st != "OK":
                for part in st.split(","):
                    p = part.strip()
                    if p and p not in issues: issues.append(p)
            mods.append({"name": unescape(m.get("name","")), "status": st,
                "platform": unescape(m.get("platform","")),
                "size": unescape(m.get("size","")), "md5": m.get("checksum",""),
                "has_fatal": m.get("has_fatal_errors","false").lower() == "true",
                "is_dep": m.get("is_dependency","false").lower() == "true",
                "issues": issues, "is_selected": False, "is_ignored": False, "is_third_party": False})
        return mods

    def get_app_info(self, aid: int) -> dict | None:
        try: root = self._xml("getappinfo.do", {"app_id": aid})
        except (requests.HTTPError, ET.ParseError): return None
        a = root.find(".//application")
        return {"modified": a.get("modified_date","")} if a is not None else None


# ==========================================================================
# Module merging
# ==========================================================================

def _merge_modules(dr_modules: list[dict], prescan_modules: list[dict]) -> list[dict]:
    by_name: dict[str, dict] = {}
    for m in prescan_modules:
        n = m["name"]
        if n not in by_name:
            by_name[n] = {**m, "was_scanned": False, "compiler": "", "os": "", "arch": ""}
        else:
            ex = by_name[n]
            if m["has_fatal"]: ex["has_fatal"] = True
            for iss in m["issues"]:
                if iss not in ex["issues"]: ex["issues"].append(iss)
    for dm in dr_modules:
        n = dm["name"]
        if n in by_name:
            by_name[n]["is_selected"] = True; by_name[n]["was_scanned"] = True
            by_name[n]["compiler"] = dm.get("compiler","")
            by_name[n]["os"] = dm.get("os",""); by_name[n]["arch"] = dm.get("arch","")
        else:
            by_name[n] = {"name": n, "status": "OK", "platform": "", "size": "",
                "md5": "", "has_fatal": False, "is_dep": False, "issues": [],
                "is_selected": True, "was_scanned": True, "is_ignored": False,
                "is_third_party": False, "compiler": dm.get("compiler",""),
                "os": dm.get("os",""), "arch": dm.get("arch","")}
    return list(by_name.values())


# ==========================================================================
# Individual check functions
# ==========================================================================

CheckFunc = Callable[[list[dict], list[dict], FlawSummary, dict, bool, list[str], str], tuple[list[Issue], list[str]]]

def _chk(sev: str, msg: str) -> Issue: return Issue(severity=sev, description=msg)

def check_01_junk_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; ign: list[str] = []
    for idx, f in enumerate(files):
        nm = f["name"]
        if nm.lower().endswith(".pdb") or nm.lower() in (".gitignore","head"):
            files[idx]["is_ignored"] = True; continue
        if _fancy_match(nm, JUNK_FILE_PATTERNS):
            files[idx]["is_ignored"] = True; ign.append(nm)
    for idx, m in enumerate(modules):
        if _fancy_match(m["name"], JUNK_FILE_PATTERNS): modules[idx]["is_ignored"] = True
    if ign:
        msg = f'An unnecessary file was uploaded: "{ign[0]}".' if len(ign)==1 else f'{len(ign)} unnecessary files were uploaded: {_top5(ign)}.'
        issues.append(_chk("medium", msg))
        recs.append("Follow the packaging instructions or use the Veracode auto-packager (https://docs.veracode.com/r/About_auto_packaging) to keep the upload as small as possible to improve upload and scan times.")
    return issues, recs

def check_02_third_party(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, sca_comps: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; sel_tp: list[str] = []
    for idx, f in enumerate(files):
        if _fancy_match(f["name"], THIRD_PARTY_PATTERNS): files[idx]["is_third_party"] = True
    for idx, f in enumerate(files):
        if not f["is_third_party"] and f["name"] in sca_comps: files[idx]["is_third_party"] = True
    for idx, m in enumerate(modules):
        if _fancy_match(m["name"], THIRD_PARTY_PATTERNS):
            modules[idx]["is_third_party"] = True
            if m.get("is_selected"): sel_tp.append(m["name"])
        if not m["is_third_party"] and m["name"] in sca_comps:
            modules[idx]["is_third_party"] = True
    if sel_tp:
        msg = f'A third-party component was selected as an entry point: "{sel_tp[0]}".' if len(sel_tp)==1 else f'{len(sel_tp)} third-party components selected as entry points: {_top5(sel_tp)}.'
        issues.append(_chk("medium", msg)); recs.append("Only select first party components as the entry points for the analysis.")
    return issues, recs

def check_03_flaw_count(files: list[dict], modules: list[dict], flaws: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if flaws.total == 0:
        issues.append(_chk("medium", "No flaws were found in this scan. This is usually due to scan misconfiguration."))
        recs.append("When no flaws have been found this can be an indication that incorrect modules were selected, or the main application was not selected for analysis.")
    elif flaws.total > MAX_FLAW_COUNT:
        issues.append(_chk("medium", "A large number of flaws were reported in this scan."))
        recs.append(f"More than {MAX_FLAW_COUNT} flaws were found which can be an indication that the scan could be misconfigured.")
    return issues, recs

def check_04_fatal_errors(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    # 4a. missing primary debug symbols (.NET)
    fp = [m["name"] for m in modules if m["has_fatal"] and _is_dotnet(m["name"]) and not m["is_ignored"] and not m["is_third_party"] and _has_status(m, "Primary Files Compiled without Debug Symbols")]
    if fp:
        msg = f'{len(fp)} module(s) could not be scanned due to missing debug symbols (PDB): {_top5(fp)}.' if len(fp)>1 else f'Module could not be scanned due to missing debug symbols (PDB): "{fp[0]}".'
        issues.append(_chk("high", msg)); recs.append("Include PDB files for as many components as possible, especially first and second party components.")
    # 4b. no scannable Java binaries
    fj = [m["name"] for m in modules if m["has_fatal"] and _is_java(m["name"]) and _has_status(m, "No Scannable Binaries")]
    if fj:
        msg = f'{len(fj)} Java module(s) contained no compiled Java classes: {_top5(fj)}.' if len(fj)>1 else f'Java module contained no compiled classes: "{fj[0]}".'
        issues.append(_chk("high", msg)); recs.append("Veracode requires Java apps compiled into JAR, WAR or EAR.")
    # 4c. nested JARs
    fn = [m["name"] for m in modules if m["has_fatal"] and _is_java(m["name"]) and _has_status(m, "does not support jar files nested inside")]
    if fn:
        msg = f'{len(fn)} Java module(s) contained nested/shaded JARs: {_top5(fn)}.' if len(fn)>1 else f'Java module contained nested JARs: "{fn[0]}".'
        issues.append(_chk("high", msg)); recs.append("Veracode does not support nested JARs except for Spring Boot.")
    return issues, recs

def check_05_unscannable_java(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    uj = [m["name"] for m in modules if _is_java(m["name"]) and m["has_fatal"]
          and not _has_status(m, "No Scannable Binaries") and not _has_status(m, "does not support jar files nested inside")]
    if uj:
        issues.append(_chk("high", f'{len(uj)} Java module(s) not scannable: {_top5(uj)}.'))
        recs.append("Veracode requires Java apps compiled into JAR, WAR or EAR.")
        recs.append("The Veracode CLI can be used to package Java apps: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_06_unwanted_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    for pats, ftype, rr in [
        (UNWANTED_7Z,"7-zip file",["Veracode does not support 7-zip. Consider zip files instead."]),
        (UNWANTED_COFFEE,"CoffeeScript file",["CoffeeScript is not supported.","Review JS/TS packaging: https://docs.veracode.com/r/compilation_jscript."]),
        (UNWANTED_SCRIPTS,"batch/shell script",["Do not upload batch/shell scripts."]),
        (UNWANTED_INSTALLERS,"installer",["Do not upload installers or setup programs."]),
        (UNWANTED_PYD,"Python-compiled DLL",["Do not upload .pyd files."]),
        (UNWANTED_PYC,"compiled Python file",["Veracode requires Python source code. Do not upload compiled .pyc."]),
        (UNWANTED_DEPLOY,'ClickOnce ".deploy" file',["Veracode does not support ClickOnce deployments."]),
        (UNWANTED_WIBU,"CodeMeter obfuscation file",["Do not use code obfuscation tools other than Dotfuscator Community Edition."]),
    ]:
        ff = _fancy_match_files(files, pats)
        if ff:
            msg = f'{len(ff)} {ftype}(s) uploaded: {_top5(ff)}.' if len(ff)>1 else f'A {ftype} was uploaded: "{ff[0]}".'
            issues.append(_chk("medium", msg))
            for rec in rr: recs.append(rec)
            recs.append("Follow packaging instructions: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_07_nested_archives(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    na = list(dict.fromkeys([f["name"] for f in files if f["status"] == "Archive File Within Another Archive" and not f.get("is_ignored")]))
    if na:
        issues.append(_chk("high", f'{len(na)} nested archive(s) uploaded: {_top5(na)}. Veracode does not process nested archives.'))
        recs.append("Do not upload nested archives.")
    return issues, recs

def check_08_missing_precompiled(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    pcf = _fancy_match_files(files, DOTNET_PRECOMPILE_PATTERNS)
    if pcf:
        issues.append(_chk("high", f'{len(pcf)} .NET view/template file(s) uploaded: {_top5(pcf)}. Precompile ASP.NET views.'))
        recs.append("Precompile ASP.NET projects and upload all generated assemblies.")
    pc_mods = [m["name"] for m in modules if m.get("is_selected") and _is_dotnet(m["name"]) and not m["is_ignored"] and not m["is_third_party"] and any("No precompiled files were found" in iss for iss in m.get("issues",[]))]
    if pc_mods:
        issues.append(_chk("medium", f'{len(pc_mods)} .NET component(s) missing precompiled files: {_top5(pc_mods)}.'))
    return issues, recs

def check_09_missing_sca(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, sca_on: bool, sca_comps: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if sca_on and not sca_comps:
        if any(_fancy_match(f["name"], SCA_SUPPORTED) for f in files):
            issues.append(_chk("medium", "No SCA results for this scan. Possible misconfiguration."))
            recs.append("Follow packaging guidance: https://docs.veracode.com/r/compilation_packaging.")
    return issues, recs

def check_10_unselected_js(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    uj = list(dict.fromkeys([m["name"] for m in modules if _is_js_module(m["name"]) and not _is_node_module(m["name"])
              and "extracted from .map file" not in m["name"].lower()
              and not m["has_fatal"] and not m["is_ignored"] and not m.get("is_selected") and not m["is_third_party"]]))
    if uj:
        issues.append(_chk("medium", f'{len(uj)} JS module(s) not selected: {_top5(uj)}.'))
        recs.append('Select "JS files within ..." modules for JavaScript coverage.')
        recs.append("Under-selection of first party modules affects results quality: https://community.veracode.com/s/article/What-are-Modules-and-how-do-my-results-change-based-on-what-I-select.")
    return issues, recs

def check_11_unexpected_source(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    for pats, ft, rr in [
        (SRC_JAVA,"Java source",["Do not upload Java source. Compile into JAR/WAR/EAR: https://docs.veracode.com/r/compilation_java."]),
        (SRC_CS,"C# source",["Do not upload C# source. Compile with debug symbols: https://docs.veracode.com/r/compilation_net."]),
        (SRC_SLN,".NET solution file",["Do not upload .sln files."]),
        (SRC_CSPROJ,"C# project file",["Do not upload .csproj files."]),
        (SRC_C,"C source",["Do not upload C source. Compile with debug symbols."]),
        (SRC_CPP,"C++ source",["Do not upload C++ source. Compile with debug symbols."]),
        (SRC_SWIFT,"Swift source",["Do not upload Swift source. Compile per iOS packaging guidelines."]),
    ]:
        ff = _fancy_match_files(files, pats)
        if ff:
            issues.append(_chk("high", f'{len(ff)} {ft} file(s) uploaded: {_top5(ff)}. Veracode requires compiled binaries.'))
            for rec in rr: recs.append(rec)
    return issues, recs

def check_12_missing_supporting(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; ms_mods: list[str] = []; ms_count = 0
    for m in modules:
        if not m.get("is_selected") or m["is_ignored"]: continue
        for iss in m.get("issues",[]):
            if iss.startswith("Missing Supporting Files"):
                parts = iss.split(" ")
                if len(parts) > 4:
                    try: ms_count += int(parts[4]); ms_mods.append(m["name"]) if m["name"] not in ms_mods else None
                    except ValueError: pass
    if ms_count:
        issues.append(_chk("medium", f'{len(ms_mods)} module(s) missing {ms_count} supporting file(s): {_top5(ms_mods)}.'))
        recs.append("Resolve missing supporting files on the Review Modules page.")
        recs.append("Ensure all components are present for analysis.")
    return issues, recs

def check_13_missing_debug(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    md = list(dict.fromkeys([m["name"] for m in modules if _is_dotnet(m["name"]) and not m["is_ignored"] and not m["is_third_party"]
          and any("No supporting files or PDB files" in iss for iss in m.get("issues",[]))]))
    if md:
        issues.append(_chk("medium", f'{len(md)} module(s) lack debug symbols (PDB): {_top5(md)}.'))
        recs.append("Include PDB files for first and second party components.")
    return issues, recs

def check_14_unsupported_platform(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    upc = list(dict.fromkeys([m["name"] for m in modules if m["has_fatal"] and not m["is_ignored"] and not m["is_third_party"]
           and (_has_status(m,"(Fatal)Unsupported Platform") or _has_status(m,"(Fatal)Unsupported Compiler"))]))
    if upc:
        issues.append(_chk("high", f'{len(upc)} module(s) have unsupported platform/compiler: {_top5(upc)}.'))
        recs.append("Review packaging docs to ensure compiler is supported.")
    return issues, recs

def check_15_gradle_wrapper(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    gw = _fancy_match_modules(modules, GRADLE_WRAPPER, selected_only=True)
    if gw:
        issues.append(_chk("high", '"gradle-wrapper.jar" selected for analysis. This is a build tool, not the application.'))
        recs.append('Do not upload or select "gradle-wrapper.jar".')
        recs.append("Use Veracode CLI: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_16_sensitive_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    for pats, desc, sr in [
        (SENSITIVE_SECRET_PATTERNS, "potentially sensitive/secret file", ["Do not upload secrets, certificates, or keys."]),
        (SENSITIVE_BACKUP_PATTERNS, "backup/old file", ["Do not upload backup files."]),
        (SENSITIVE_WORD_PATTERNS, "Word document", ["Office documents could contain sensitive information."]),
        (SENSITIVE_SPREADSHEET_PATTERNS, "spreadsheet", ["Office documents could contain sensitive information."]),
        (SENSITIVE_JUPYTER_PATTERNS, "Jupyter notebook", ["Jupyter notebooks could contain sensitive data."]),
    ]:
        ff = _fancy_match_files(files, pats)
        if ff:
            issues.append(_chk("high", f'{len(ff)} {desc}(s) uploaded: {_top5(ff)}.'))
            for rec in sr: recs.append(rec)
            recs.append("Do not upload unnecessary files.")
    return issues, recs

def check_17_repositories(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    repo_f = [f["name"] for f in files if f["name"].lower() in ("fsmonitor-watchman.sample","fetch_head")]
    if repo_f:
        issues.append(_chk("medium", "A git repository was uploaded. Repositories can contain sensitive information."))
        recs.append("Do not upload source code repositories.")
    return issues, recs

def check_18_node_modules(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []; nm_found: list[str] = []
    for idx, m in enumerate(modules):
        if _is_node_module(m["name"]):
            if m["name"] not in nm_found: nm_found.append(m["name"])
            modules[idx]["is_third_party"] = True; modules[idx]["is_ignored"] = True
    if nm_found:
        issues.append(_chk("medium", f'{len(nm_found)} "node_modules" folder(s) uploaded. This increases upload size and module count.'))
        recs.append('Do not upload "node_modules" folders.')
        recs.append("Use Veracode CLI: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_19_testing_artefacts(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    sel_test = _fancy_match_modules(modules, TEST_FILE_PATTERNS, selected_only=True)
    if sel_test:
        issues.append(_chk("high", f'{len(sel_test)} testing artefact(s) selected: {_top5(sel_test)}.'))
        recs.append("Do not upload testing artifacts. Do not select them as entry points.")
    upl_test = _fancy_match_files(files, TEST_FILE_PATTERNS)
    if upl_test:
        issues.append(_chk("medium", f'{len(upl_test)} testing artefact(s) uploaded: {_top5(upl_test)}.'))
        recs.append("Do not upload testing artifacts.")
    test_mods = list(dict.fromkeys([m["name"] for m in modules for iss in m.get("issues",[]) if "test/" in iss.lower()]))
    if test_mods:
        issues.append(_chk("medium", f'{len(test_mods)} module(s) contain testing artefacts: {_top5(test_mods)}.'))
    for idx, f in enumerate(files):
        if _fancy_match(f["name"], TEST_FILE_PATTERNS): files[idx]["is_ignored"] = True
    for idx, m in enumerate(modules):
        if _fancy_match(m["name"], TEST_FILE_PATTERNS): modules[idx]["is_ignored"] = True
    return issues, recs

def check_20_too_many_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if len(files) > MAX_FILE_COUNT:
        issues.append(_chk("medium", f"Too many files uploaded ({len(files)}). May cause many modules and long scan times."))
    return issues, recs

def check_21_excess_microsoft(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    ef = _fancy_match_files(files, EXCESS_MSFT_PATTERNS)
    if ef:
        issues.append(_chk("medium", f'{len(ef)} .NET Roslyn/Runtime component(s) uploaded: {_top5(ef)}.'))
        recs.append("Do not include unnecessary Microsoft runtime components.")
    return issues, recs

def check_22_loose_class_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    lcf = _fancy_match_files(files, LOOSE_CLASS_FILE)
    lcm = _fancy_match_modules(modules, LOOSE_CLASS_MODULE)
    if lcf or lcm:
        issues.append(_chk("medium", "Java class files not packaged in JAR/WAR/EAR. Suboptimal compilation."))
        recs.append("Compile Java into JAR/WAR/EAR per packaging instructions.")
        recs.append("Use Veracode CLI: https://docs.veracode.com/r/About_auto_packaging.")
    return issues, recs

def check_23_go_workspace(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    gwf = _fancy_match_files(files, GO_WORKSPACE_PATTERNS)
    if gwf:
        issues.append(_chk("medium", "Go workspaces detected. Multi-module workspaces not supported."))
        recs.append("Follow Go packaging: https://docs.veracode.com/r/compilation_go.")
    return issues, recs

def check_24_unselected_first_party(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    ufp = list(dict.fromkeys([m["name"] for m in modules if not m["is_dep"] and not m["is_ignored"]
           and not m.get("is_selected") and not m["is_third_party"] and not _is_js_module(m["name"]) and not m["has_fatal"]]))
    if ufp:
        issues.append(_chk("medium", f'{len(ufp)} first-party module(s) not selected: {_top5(ufp)}.'))
        recs.append("Under-selection affects quality: https://community.veracode.com/s/article/What-are-Modules-and-how-do-my-results-change-based-on-what-I-select.")
    return issues, recs

def check_25_over_scanning(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    """Detect modules selected that are dependencies of other selected modules."""
    issues: list[Issue] = []; recs: list[str] = []
    selected = {m["name"] for m in modules if m.get("is_selected") and not m["is_ignored"]}
    dep_names = {m["name"] for m in modules if m["is_dep"]}
    over = sorted(selected & dep_names)
    if over:
        issues.append(_chk("medium", f'{len(over)} module(s) selected but are dependencies of other modules: {_top5(over)}. This can lead to duplicate flaw reporting.'))
        recs.append("Only select main entry points, not dependency libraries.")
    return issues, recs

def check_26_dependencies_selected(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    dep_sel = list(dict.fromkeys([m["name"] for m in modules if m.get("is_selected") and not m["is_ignored"] and m["is_dep"]]))
    if dep_sel:
        issues.append(_chk("medium", f'{len(dep_sel)} dependenc{"y" if len(dep_sel)==1 else "ies"} selected as entry point(s): {_top5(dep_sel)}.'))
        recs.append("Only select main entry points: https://community.veracode.com/s/article/What-are-Modules-and-how-do-my-results-change-based-on-what-I-select.")
    return issues, recs

def check_27_duplicate_files(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    seen: dict[str, list[str]] = {}
    for f in files:
        if f.get("is_ignored") or f.get("is_third_party"): continue
        seen.setdefault(f["name"], []).append(f.get("md5",""))
    diff = {n: h for n, h in seen.items() if len(h) > 1 and len(set(h)) > 1}
    same = {n: len(h) for n, h in seen.items() if len(h) > 1 and len(set(h)) == 1}
    if diff:
        issues.append(_chk("high", f'{len(diff)} duplicate filename(s) with different hashes: {_top5(list(diff.keys()))}. Can cause indeterministic results.'))
    if same:
        issues.append(_chk("medium", f'{len(same)} duplicate file(s) uploaded. Slows scan time.'))
    if diff or same:
        recs.append("De-duplicate modules before upload.")
        recs.append("Upload only one version of each component.")
    return issues, recs

def check_28_minified_js(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    mf = list(dict.fromkeys([f["name"] for f in files if f["name"].lower().endswith(".min.js")]))
    if mf:
        issues.append(_chk("medium", f'{len(mf)} minified JS file(s) uploaded: {_top5(mf)}. Will not be scanned.'))
        recs.append("Submit readable JavaScript source: https://docs.veracode.com/r/compilation_jscript.")
    mm = list(dict.fromkeys([m["name"] for m in modules if _is_js_module(m["name"])
        for iss in m.get("issues",[]) if "because we think it is minified" in iss or "dist/" in iss.lower()]))
    if mm:
        issues.append(_chk("medium", f'{len(mm)} minified JS within modules: {_top5(mm)}.'))
    return issues, recs

def check_29_module_count(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    sel = [m for m in modules if m.get("is_selected")]
    if len(sel) > MAX_SELECTED_MODULE_COUNT:
        issues.append(_chk("medium", f"{len(sel)} modules selected. Suboptimal configuration."))
        recs.append("Select correct modules. Consider splitting application profiles.")
    if len(modules) > MAX_MODULE_COUNT:
        issues.append(_chk("medium", f"{len(modules)} modules identified. Suboptimal upload."))
        recs.append("Follow packaging guidance.")
    return issues, recs

def check_30_regular_scans(files: list[dict], modules: list[dict], _f: FlawSummary, _m: dict, _s: bool, _sc: list[str], app_mod: str) -> tuple[list[Issue], list[str]]:
    issues: list[Issue] = []; recs: list[str] = []
    if app_mod:
        ds = _days_since(app_mod)
        if ds is not None and ds > STALE_SCAN_DAYS:
            issues.append(_chk("medium", f"Application not scanned recently (last activity {ds} days ago)."))
            recs.append("Regular scanning via automation allows faster response to new issues.")
    return issues, recs

def check_31_analysis_size(files: list[dict], modules: list[dict], _f: FlawSummary, scan_meta: dict, _s: bool, _sc: list[str], _a: str) -> tuple[list[Issue], list[str]]:
    """Flag when analysis size or total module size exceeds thresholds."""
    issues: list[Issue] = []; recs: list[str] = []
    asz = scan_meta.get("analysis_size", 0)
    if asz > MAX_ANALYSIS_SIZE:
        issues.append(_chk("medium", f"Analysis size ({asz/(1024*1024):.0f} MB) exceeds {MAX_ANALYSIS_SIZE//(1024*1024)} MB threshold."))
        recs.append("Review packaging to exclude unnecessary files and third-party libraries.")
    total_mod_size = sum(_parse_module_size(m.get("size","")) for m in modules)
    if total_mod_size > MAX_TOTAL_MODULE_SIZE:
        issues.append(_chk("medium", f"Total module size ({total_mod_size/(1024*1024):.0f} MB) exceeds {MAX_TOTAL_MODULE_SIZE//(1024*1024)} MB threshold."))
        recs.append("Reduce upload size by following packaging instructions.")
    return issues, recs


CHECK_REGISTRY: list[tuple[int, str, CheckFunc]] = [
    (1, "ignoreJunkFiles", check_01_junk_files),
    (2, "thirdParty", check_02_third_party),
    (3, "flawCount", check_03_flaw_count),
    (4, "fatalErrors", check_04_fatal_errors),
    (5, "unscannableJava", check_05_unscannable_java),
    (6, "detectUnwantedFiles", check_06_unwanted_files),
    (7, "nestedArchives", check_07_nested_archives),
    (8, "missingPrecompiled", check_08_missing_precompiled),
    (9, "missingSCA", check_09_missing_sca),
    (10, "unselectedJS", check_10_unselected_js),
    (11, "unexpectedSource", check_11_unexpected_source),
    (12, "missingSupporting", check_12_missing_supporting),
    (13, "missingDebug", check_13_missing_debug),
    (14, "unsupportedPlatform", check_14_unsupported_platform),
    (15, "gradleWrapper", check_15_gradle_wrapper),
    (16, "sensitiveFiles", check_16_sensitive_files),
    (17, "repositories", check_17_repositories),
    (18, "nodeModules", check_18_node_modules),
    (19, "testingArtefacts", check_19_testing_artefacts),
    (20, "tooManyFiles", check_20_too_many_files),
    (21, "excessMicrosoft", check_21_excess_microsoft),
    (22, "looseClassFiles", check_22_loose_class_files),
    (23, "goWorkspace", check_23_go_workspace),
    (24, "unselectedFirstParty", check_24_unselected_first_party),
    (25, "overScanning", check_25_over_scanning),
    (26, "dependenciesSelected", check_26_dependencies_selected),
    (27, "duplicateFiles", check_27_duplicate_files),
    (28, "minifiedJS", check_28_minified_js),
    (29, "moduleCount", check_29_module_count),
    (30, "regularScans", check_30_regular_scans),
    (31, "analysisSize", check_31_analysis_size),
]

def run_checks(files: list[dict], modules: list[dict], flaws: FlawSummary,
               scan_meta: dict, sca_on: bool, sca_comps: list[str],
               app_mod: str, skip: set[int] | None = None
               ) -> tuple[list[Issue], list[str], dict[int, list[str]]]:
    """Returns (all_issues, all_recs, check_recs) where check_recs maps check# to its recs."""
    all_issues: list[Issue] = []; all_recs: list[str] = []
    check_recs: dict[int, list[str]] = {}
    for num, name, func in CHECK_REGISTRY:
        if skip and num in skip: continue
        try:
            iss, rcs = func(files, modules, flaws, scan_meta, sca_on, sca_comps, app_mod)
            for i in iss:
                i.check_num = num
                i.check_name = name
            all_issues.extend(iss)
            if rcs:
                check_recs[num] = rcs
            for r in rcs:
                if r not in all_recs: all_recs.append(r)
        except Exception as e:
            log.warning("Check #%d (%s) failed: %s", num, name, e)
    return all_issues, all_recs, check_recs


# ==========================================================================
# Orchestration
# ==========================================================================

def _find_latest_published_build(client: VeracodeClient, aid: int, builds: list[dict]) -> dict | None:
    """Return the latest build that has published results, searching from newest to oldest."""
    for b in reversed(builds):
        bi = client.get_build_info(aid, b["id"])
        if bi and bi.get("published"):
            return b
    # Fallback: return the last build even if not published
    return builds[-1] if builds else None


def _process_build(client: VeracodeClient, app: dict, builds: list[dict],
                   legacy_id: int, sandbox_name: str = "",
                   skip_checks: set[int] | None = None,
                   prev_data: dict | None = None) -> tuple[ScanResult, list[ModuleRow], list[FileRow], list[RecommendationRow], list[AggIssue], TrendRow | None]:
    build = _find_latest_published_build(client, legacy_id, builds)
    if build is None:
        sr = _empty_result(app, legacy_id, sandbox_name)
        sr.issues_text = "[HIGH] No published build found"
        sr.health = "Poor"; sr.high_issues = 1; sr.total_issues = 1
        return sr, [], [], [], [], None

    bid = build["id"]
    dr = client.get_detailed_report(bid)
    files = client.get_files(legacy_id, bid)
    prescan = client.get_prescan(legacy_id, bid)
    app_info = client.get_app_info(legacy_id)
    app_mod_date = (app_info or {}).get("modified","") if not sandbox_name else ""

    if dr is None:
        fl = FlawSummary()
        dr = {"account_id":"","app_id":"","sandbox_id":0,"sandbox_name":sandbox_name,
              "analysis_id":"","sau_id":"","bu":"","app_name":app["name"],
              "scan_name":build.get("ver",""),"engine":"",
              "submitted":"","published":"","analysis_size":0,
              "is_latest":True,"flaws":fl,"dr_modules":[],"sca_on":False,"sca_comps":[]}
    else:
        fl = dr["flaws"]

    modules = _merge_modules(dr["dr_modules"], prescan)

    base = client.base
    acct = dr.get("account_id",""); aid_str = dr.get("app_id","")
    an_id = dr.get("analysis_id",""); sau = dr.get("sau_id",""); sbx_id = dr.get("sandbox_id",0)
    rev_url = f"{base}/auth/index.jsp#AnalyzeAppModuleList:{acct}:{aid_str}:{bid}:{an_id}:{sau}::::{sbx_id}" if acct else ""
    tri_url = f"{base}/auth/index.jsp#ReviewResultsStaticFlaws:{acct}:{aid_str}:{bid}:{an_id}:{sau}::::{sbx_id}" if acct else ""

    scan_meta = {"review_url": rev_url, "triage_url": tri_url, "analysis_size": dr.get("analysis_size",0)}
    issues, recs, check_recs = run_checks(files, modules, fl, scan_meta,
                              dr.get("sca_on",False), dr.get("sca_comps",[]),
                              app_mod_date, skip_checks)

    hi = sum(1 for i in issues if i.severity=="high")
    mi = sum(1 for i in issues if i.severity=="medium")
    health = "Good" if not issues else ("Poor" if hi else ("Fair" if mi else "Good"))

    ds = _days_since(dr.get("published",""))
    dur = _dur(dr.get("submitted",""), dr.get("published",""))
    sel = [m for m in modules if m.get("is_selected")]
    sel_names = ", ".join(m["name"] for m in sel)
    total_upload = sum(_parse_module_size(m.get("size","")) for m in modules)
    asz = dr.get("analysis_size",0)

    # Trend
    trend: TrendRow | None = None
    trend_label = ""
    key = (app["name"], sandbox_name or dr.get("sandbox_name",""))
    if prev_data and key in prev_data:
        prev = prev_data[key]
        ph = prev.get("Health",""); pf = _si(prev.get("Total Flaws")); po = _si(prev.get("Open Affecting Policy"))
        if ph and health != ph:
            trend_label = "Improved" if (health == "Good" or (health == "Fair" and ph == "Poor")) else "Degraded"
        elif ph: trend_label = "Unchanged"
        else: trend_label = "New"
        trend = TrendRow(app_name=app["name"], sandbox=sandbox_name or dr.get("sandbox_name",""),
            prev_health=ph, curr_health=health, change=trend_label,
            prev_flaws=pf, curr_flaws=fl.total, flaw_delta=fl.total-pf,
            prev_open_pol=po, curr_open_pol=fl.open_pol, open_pol_delta=fl.open_pol-po)
    elif prev_data:
        trend_label = "New"

    sr = ScanResult(
        app_name=app["name"], bu=app.get("bu",""), policy=app.get("policy",""),
        app_id=legacy_id, sandbox=sandbox_name or dr.get("sandbox_name",""),
        build_id=bid, scan_name=dr.get("scan_name",""),
        is_latest=dr.get("is_latest",True),
        scan_status="Results Ready" if dr.get("published") else "No Results",
        published=dr.get("published",""), days_since=ds if ds is not None else "N/A",
        duration=dur, engine=dr.get("engine",""),
        analysis_size_mb=round(asz/(1024*1024),2) if asz else 0,
        total_upload_mb=round(total_upload/(1024*1024),2) if total_upload else 0,
        files_uploaded=len(files), total_modules=len(modules),
        selected_modules=len(sel), selected_names=sel_names[:500],
        fatal_errors=sum(1 for m in modules if m["has_fatal"]),
        flaws=fl, sca_count=len(dr.get("sca_comps",[])),
        age_bucket=_age_bucket(ds),
        health=health, health_trend=trend_label,
        high_issues=hi, medium_issues=mi, total_issues=len(issues),
        issues_text="; ".join(f"[{i.severity.upper()}] {i.description}" for i in issues) if issues else "None",
        recs_text="; ".join(recs) if recs else "None",
        review_url=rev_url, triage_url=tri_url,
    )

    mod_rows = [ModuleRow(app_name=app["name"], build_id=bid, name=m["name"],
                status=m.get("status",""), selected=m.get("is_selected",False),
                dependency=m["is_dep"], fatal=m["has_fatal"],
                third_party=m["is_third_party"], ignored=m["is_ignored"],
                platform=m.get("platform",""), compiler=m.get("compiler",""),
                size=m.get("size",""), issues="; ".join(m.get("issues",[])))
                for m in modules]

    file_rows = [FileRow(app_name=app["name"], build_id=bid, name=f["name"],
                 status=f["status"], md5=f["md5"],
                 ignored=f["is_ignored"], third_party=f["is_third_party"])
                 for f in files]

    # Build recommendation rows -- match each rec to the highest-severity issue it came from
    rec_rows: list[RecommendationRow] = []
    for r in recs:
        # Find the most severe issue whose check produced this rec
        best_sev = "low"
        for i in issues:
            if best_sev != "high" and i.severity == "high": best_sev = "high"
            elif best_sev == "low" and i.severity == "medium": best_sev = "medium"
        rec_rows.append(RecommendationRow(app_name=app["name"], build_id=bid,
                    severity=best_sev, recommendation=r, doc_url=_extract_url(r)))

    # Build structured aggregation issues (one per issue, with check metadata and matched rec)
    sb_name = sandbox_name or dr.get("sandbox_name", "")
    agg_issues: list[AggIssue] = []
    for i in issues:
        # Get the first recommendation from this specific check
        check_rec_list = check_recs.get(i.check_num, [])
        matched_rec = check_rec_list[0] if check_rec_list else ""
        agg_issues.append(AggIssue(
            app_name=app["name"], bu=app.get("bu", ""), sandbox=sb_name,
            check_num=i.check_num, check_name=i.check_name,
            category=CHECK_CATEGORIES.get(i.check_num, "Other"),
            severity=i.severity, description=i.description,
            recommendation=matched_rec,
        ))

    return sr, mod_rows, file_rows, rec_rows, agg_issues, trend


def _empty_result(app: dict, lid: int, sandbox: str = "") -> ScanResult:
    return ScanResult(app_name=app["name"], bu=app.get("bu",""), policy=app.get("policy",""),
        app_id=lid, sandbox=sandbox, scan_status="No Scan", health="Poor",
        high_issues=1, total_issues=1, issues_text="[HIGH] No policy scan found",
        recs_text="None", age_bucket="N/A")


def _process_app(client: VeracodeClient, app: dict, skip_no: bool, inc_sb: bool,
                 skip_checks: set[int] | None, prev_data: dict | None,
                 resume_keys: set[tuple[str, str]] | None
                 ) -> tuple[list[ScanResult], list[ModuleRow], list[FileRow], list[RecommendationRow], list[AggIssue], list[TrendRow]]:
    lid = app.get("legacy_id")
    if not lid: return [], [], [], [], [], []
    rs: list[ScanResult] = []; ms: list[ModuleRow] = []; fs: list[FileRow] = []
    rrs: list[RecommendationRow] = []; ais: list[AggIssue] = []; ts: list[TrendRow] = []

    key = (app["name"], "")
    if resume_keys and key in resume_keys:
        log.debug("    Skipping (resume): %s", app["name"]); return rs, ms, fs, rrs, ais, ts

    builds = client.get_builds(lid)
    if not builds:
        if skip_no: return rs, ms, fs, rrs, ais, ts
        rs.append(_empty_result(app, lid))
        return rs, ms, fs, rrs, ais, ts

    sr, mr, fr, rr, ai, tr = _process_build(client, app, builds, lid,
                                         skip_checks=skip_checks, prev_data=prev_data)
    rs.append(sr); ms.extend(mr); fs.extend(fr); rrs.extend(rr); ais.extend(ai)
    if tr: ts.append(tr)

    if inc_sb:
        for sb in client.get_sandboxes(lid):
            sb_key = (app["name"], sb["name"])
            if resume_keys and sb_key in resume_keys: continue
            sb_builds = client.get_builds(lid, sb["id"])
            if not sb_builds: continue
            sr2, mr2, fr2, rr2, ai2, tr2 = _process_build(client, app, sb_builds, lid, sb["name"],
                                                       skip_checks=skip_checks, prev_data=prev_data)
            rs.append(sr2); ms.extend(mr2); fs.extend(fr2); rrs.extend(rr2); ais.extend(ai2)
            if tr2: ts.append(tr2)

    return rs, ms, fs, rrs, ais, ts


# ==========================================================================
# Resume
# ==========================================================================

def _load_resume_keys(path: str) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb["Scan Health Summary"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        ai = headers.index("App Name") if "App Name" in headers else -1
        si = headers.index("Sandbox") if "Sandbox" in headers else -1
        if ai >= 0:
            for row in ws.iter_rows(min_row=2, values_only=True):
                keys.add((str(row[ai] or ""), str(row[si] or "") if si >= 0 else ""))
        wb.close()
    except Exception as e:
        log.warning("Could not load resume file: %s", e)
    return keys


# ==========================================================================
# Trend
# ==========================================================================

def _load_previous(path: str) -> dict[tuple[str, str], dict]:
    data: dict[tuple[str, str], dict] = {}
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb["Scan Health Summary"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rd = dict(zip(headers, row))
            key = (str(rd.get("App Name","")), str(rd.get("Sandbox","") or ""))
            data[key] = rd
        wb.close()
    except Exception as e:
        log.warning("Could not load previous report: %s", e)
    return data


# ==========================================================================
# Output
# ==========================================================================

_CW: dict[str, int] = {"App Name":30,"Issues":80,"Recommendations":80,"Module":30,
    "File":35,"Review Modules URL":50,"Triage Flaws URL":50,"Policy":25,
    "Business Unit":20,"Selected Module Names":40,"Recommendation":60,"Doc URL":40,
    "Check Name":22,"Category":18,"Issue Pattern":60,"Top Recommendation":60,
    "Affected App Names":50,"Business Units":35}

def _hdr(ws: object, n: int) -> None:
    for c in range(1,n+1):
        cl=ws.cell(row=1,column=c); cl.fill=_HF; cl.font=_HN; cl.alignment=_CA; cl.border=_BD

def _sheet(ws: object, rows: list[dict], hcol: str | None = None, age_col: str | None = None) -> None:
    if not rows: return
    hds=list(rows[0].keys()); ws.append(hds); _hdr(ws,len(hds)); ws.freeze_panes="A2"
    for ri,rd in enumerate(rows,2):
        for ci,h in enumerate(hds,1):
            cl=ws.cell(row=ri,column=ci,value=rd.get(h,"")); cl.font=_DF; cl.border=_BD
            if hcol and h==hcol:
                cl.fill=PatternFill("solid",fgColor=_CLR.get(str(rd.get(h,"")),"FFFFFF")); cl.alignment=_CA
            elif age_col and h==age_col:
                cl.fill=PatternFill("solid",fgColor=_AGE_CLR.get(str(rd.get(h,"")),"FFFFFF")); cl.alignment=_CA
            elif h in ("Issues","Recommendations","Recommendation"): cl.alignment=_WA
            else: cl.alignment=_CA
    for ci,h in enumerate(hds,1):
        ws.column_dimensions[get_column_letter(ci)].width=_CW.get(h,14)
    ws.auto_filter.ref=ws.dimensions


def _build_aggregation(agg_issues: list[AggIssue], total_apps: int) -> list[dict]:
    """Build accurate tenant-level aggregation from structured AggIssue records.

    Groups issues by (check_num, check_name) so each row represents one type of
    problem across the tenant. Uses the real check metadata rather than regex
    guesswork on serialized strings.

    Output columns:
      - Check #: the check number for reference
      - Check Name: machine-readable check name
      - Category: Packaging / Module Selection / Fatal Errors / etc.
      - Severity: highest severity observed for this check across all apps
      - Apps Affected: count of unique apps that triggered this check
      - % of Tenant: percentage of total apps
      - Affected App Names: comma-separated, max 10 then "and N others"
      - Business Units Affected: unique BUs with this issue
      - Sample Issue: one representative issue description (shortest, for clarity)
      - Top Recommendation: the recommendation linked to this check
    """
    if not agg_issues:
        return []

    # Group by check number
    by_check: dict[int, list[AggIssue]] = {}
    for ai in agg_issues:
        by_check.setdefault(ai.check_num, []).append(ai)

    rows: list[dict] = []
    for check_num in sorted(by_check.keys()):
        group = by_check[check_num]
        check_name = group[0].check_name
        category = group[0].category

        # Highest severity in this group
        sevs = {ai.severity for ai in group}
        if "high" in sevs:
            sev = "HIGH"
        elif "medium" in sevs:
            sev = "MEDIUM"
        else:
            sev = "LOW"

        # Unique apps
        unique_apps = list(dict.fromkeys(ai.app_name for ai in group))
        app_count = len(unique_apps)
        pct = round(100 * app_count / total_apps, 1) if total_apps else 0

        app_str = ", ".join(unique_apps[:10])
        if len(unique_apps) > 10:
            app_str += f" and {len(unique_apps) - 10} others"

        # Unique business units with counts
        bu_counter: Counter[str] = Counter()
        for ai in group:
            if ai.bu: bu_counter[ai.bu] += 1
        bu_parts = [f"{bu} ({cnt})" for bu, cnt in bu_counter.most_common(10)]
        bu_str = ", ".join(bu_parts)
        if len(bu_counter) > 10:
            bu_str += f" and {len(bu_counter) - 10} others"

        # Total issue occurrences (may exceed app count when one app triggers
        # multiple sub-issues from the same check, e.g. multiple fatal error types)
        occurrence_count = len(group)

        # Representative issue: pick the most common normalized pattern
        pattern_counter: Counter[str] = Counter()
        for ai in group:
            clean = re.sub(r'"[^"]*"', '(name)', ai.description)
            clean = re.sub(r'\b\d+\b', 'N', clean)
            pattern_counter[clean] += 1
        sample = pattern_counter.most_common(1)[0][0] if pattern_counter else ""

        # Top recommendation: pick the first non-empty one
        rec = ""
        for ai in group:
            if ai.recommendation:
                rec = ai.recommendation
                break

        rows.append({
            "Check #": check_num,
            "Check Name": check_name,
            "Category": category,
            "Severity": sev,
            "Apps Affected": app_count,
            "% of Tenant": pct,
            "Total Occurrences": occurrence_count,
            "Affected App Names": app_str,
            "Business Units": bu_str,
            "Issue Pattern": sample,
            "Top Recommendation": rec,
        })

    # Sort by severity (HIGH first), then by app count descending
    sev_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    rows.sort(key=lambda r: (sev_order.get(r["Severity"], 3), -r["Apps Affected"]))

    return rows


def write_excel(health: list[dict], mods: list[dict], files: list[dict],
                recs: list[dict], trends: list[dict], agg: list[dict], path: str) -> None:
    wb=Workbook()
    ws=wb.active; ws.title="Scan Health Summary"
    if not health: ws["A1"]="No data."; wb.save(path); return
    _sheet(ws, health, hcol="Health", age_col="Scan Age Bucket")
    if mods: _sheet(wb.create_sheet("Module Details"), mods)
    if files: _sheet(wb.create_sheet("Uploaded Files"), files)
    if recs: _sheet(wb.create_sheet("Recommendations"), recs)
    if trends: _sheet(wb.create_sheet("Trends"), trends)
    if agg: _sheet(wb.create_sheet("Tenant Aggregation"), agg)
    # Overview
    wso=wb.create_sheet("Tenant Overview")
    tot=len(health); good=sum(1 for r in health if r.get("Health")=="Good")
    poor=sum(1 for r in health if r.get("Health")=="Poor")
    fair=sum(1 for r in health if r.get("Health")=="Fair")
    ns=sum(1 for r in health if r.get("Scan Status")=="No Scan")
    ds=[r["Days Since Scan"] for r in health if isinstance(r.get("Days Since Scan"),int)]
    avg=round(sum(ds)/len(ds),1) if ds else 0
    stats=[("Report Generated",datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
           ("Total Applications",tot),("Good Health",good),("Fair",fair),("Poor Health",poor),
           ("No Scan Found",ns),
           ("Total Flaws",sum(_si(r.get("Total Flaws")) for r in health)),
           ("Total Open Affecting Policy",sum(_si(r.get("Open Affecting Policy")) for r in health)),
           ("Avg Days Since Scan",avg)]
    wso.column_dimensions["A"].width=40; wso.column_dimensions["B"].width=20
    wso.append(["Metric","Value"]); _hdr(wso,2)
    for ri,(k,v) in enumerate(stats,2):
        wso.cell(row=ri,column=1,value=k).font=_BF; wso.cell(row=ri,column=2,value=v).font=_DF
        for c in (1,2): wso.cell(row=ri,column=c).border=_BD; wso.cell(row=ri,column=c).alignment=_CA
    wb.save(path); log.info("[+] Report saved: %s", path)


def write_csv(health: list[dict], mods: list[dict], files: list[dict],
              recs: list[dict], trends: list[dict], agg: list[dict], base_path: str) -> None:
    stem = Path(base_path).stem
    parent = Path(base_path).parent
    for name, rows in [("summary",health),("modules",mods),("files",files),
                       ("recommendations",recs),("trends",trends),("aggregation",agg)]:
        if not rows: continue
        p = parent / f"{stem}_{name}.csv"
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader(); w.writerows(rows)
        log.info("[+] CSV: %s", p)


def write_json(health: list[dict], mods: list[dict], files: list[dict],
               trends: list[dict], path: str) -> None:
    out = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "summary": {"total": len(health),
                     "good": sum(1 for r in health if r.get("Health")=="Good"),
                     "fair": sum(1 for r in health if r.get("Health")=="Fair"),
                     "poor": sum(1 for r in health if r.get("Health")=="Poor")},
        "apps": health, "modules": mods, "files": files,
        "trends": trends or None,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, default=str)
    log.info("[+] JSON: %s", path)


# ==========================================================================
# Self-test
# ==========================================================================

def _self_test() -> None:
    """Run checks against a mock fixture and verify expected results."""
    files = [
        {"name": "app.jar", "status": "OK", "md5": "aaa", "is_ignored": False, "is_third_party": False},
        {"name": "Thumbs.db", "status": "OK", "md5": "bbb", "is_ignored": False, "is_third_party": False},
        {"name": "secret.pem", "status": "OK", "md5": "ccc", "is_ignored": False, "is_third_party": False},
        {"name": "nested.zip", "status": "Archive File Within Another Archive", "md5": "ddd", "is_ignored": False, "is_third_party": False},
        {"name": "app.java", "status": "OK", "md5": "eee", "is_ignored": False, "is_third_party": False},
    ]
    modules = [
        {"name": "app.jar", "status": "OK", "platform": "Java", "size": "5MB", "md5": "aaa",
         "has_fatal": False, "is_dep": False, "issues": [], "is_selected": True,
         "is_ignored": False, "is_third_party": False, "compiler": "", "os": "", "arch": ""},
        {"name": "lib.jar", "status": "OK", "platform": "Java", "size": "2MB", "md5": "fff",
         "has_fatal": False, "is_dep": True, "issues": [], "is_selected": True,
         "is_ignored": False, "is_third_party": False, "compiler": "", "os": "", "arch": ""},
    ]
    flaws = FlawSummary(total=5, open_pol=2, mitigated=1, fixed=2, pol_aff=3)
    issues, recs = run_checks(files, modules, flaws, {"analysis_size": 100},
                              False, [], "2020-01-01 00:00:00 UTC")

    checks_triggered = set()
    for i in issues:
        d = i.description.lower()
        if "thumbs.db" in d or "unnecessary" in d: checks_triggered.add(1)
        if "secret" in d or "sensitive" in d or ".pem" in d: checks_triggered.add(16)
        if "nested" in d: checks_triggered.add(7)
        if "java source" in d: checks_triggered.add(11)
        if "dependency" in d or "dependenc" in d: checks_triggered.add(26)
        if "not scanned recently" in d or "not been recent" in d: checks_triggered.add(30)

    expected = {1, 7, 11, 16, 26, 30}
    missing = expected - checks_triggered
    if missing:
        print(f"FAIL: expected checks {missing} not triggered")
        print(f"Issues found: {[i.description for i in issues]}")
        raise SystemExit(1)
    print(f"PASS: {len(issues)} issues, {len(recs)} recommendations. Checks triggered: {sorted(checks_triggered)}")
    raise SystemExit(0)


# ==========================================================================
# Main
# ==========================================================================

def _print_summary(health: list[dict]) -> None:
    tot = len(health)
    good = sum(1 for r in health if r.get("Health")=="Good")
    fair = sum(1 for r in health if r.get("Health")=="Fair")
    poor = sum(1 for r in health if r.get("Health")=="Poor")
    print(f"\nTenant Summary: {tot} apps - Good: {good}, Fair: {fair}, Poor: {poor}")
    counter: Counter[str] = Counter()
    for r in health:
        txt = r.get("Issues","")
        if txt == "None": continue
        for part in txt.split("; "):
            clean = re.sub(r'\[(?:HIGH|MEDIUM|LOW)\]\s*', '', part).strip()
            clean = re.sub(r'"[^"]*"', '(name)', clean)
            if clean: counter[clean] += 1
    if counter:
        print("Top issues:")
        for pattern, count in counter.most_common(3):
            print(f"  [{count}x] {pattern[:100]}")


def main() -> None:
    p = argparse.ArgumentParser(description="Veracode Tenant-Wide Scan Health v3.0")
    p.add_argument("--output", default=f"scan_health_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    p.add_argument("--output-format", choices=["xlsx","csv","json"], default="xlsx")
    p.add_argument("--max-apps", type=int, default=0)
    p.add_argument("--delay", type=float, default=0.5)
    p.add_argument("--skip-no-scan", action="store_true")
    p.add_argument("--include-sandboxes", action="store_true")
    p.add_argument("--region", choices=["commercial","eu"], default="commercial")
    p.add_argument("--app-name-filter", default=None, help="Regex to filter app names")
    p.add_argument("--parallel", type=int, default=1, help="Concurrent workers (default 1)")
    p.add_argument("--resume", default=None, help="Path to prior partial xlsx to skip processed apps")
    p.add_argument("--previous-report", default=None, help="Path to prior xlsx for trend analysis")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--log-level", choices=["DEBUG","INFO","WARNING"], default="INFO")
    p.add_argument("--timeout", type=int, default=120)
    p.add_argument("--skip-checks", default=None, help="Comma-separated check numbers to skip")
    p.add_argument("--self-test", action="store_true")
    args = p.parse_args()

    logging.basicConfig(level=getattr(logging, args.log_level),
                        format="%(asctime)s %(levelname)-5s %(message)s", datefmt="%H:%M:%S",
                        force=True)

    if args.self_test:
        _self_test()

    skip_checks: set[int] = set()
    if args.skip_checks:
        skip_checks = {int(x.strip()) for x in args.skip_checks.split(",")}
        log.info("[*] Skipping checks: %s", sorted(skip_checks))

    resume_keys: set[tuple[str, str]] | None = None
    if args.resume:
        resume_keys = _load_resume_keys(args.resume)
        log.info("[*] Resume: %d apps already processed", len(resume_keys))

    prev_data: dict | None = None
    if args.previous_report:
        prev_data = _load_previous(args.previous_report)
        log.info("[*] Previous report: %d apps loaded for trend analysis", len(prev_data))

    name_filter = re.compile(args.app_name_filter) if args.app_name_filter else None

    try:
        with VeracodeClient(args.region, timeout=args.timeout) as client:
            log.info("[*] Region: %s", args.region)
            apps = client.get_apps()
            log.info("[*] Found %d apps", len(apps))

            if name_filter:
                apps = [a for a in apps if name_filter.search(a["name"])]
                log.info("[*] Filtered to %d apps", len(apps))
            if args.max_apps:
                apps = apps[:args.max_apps]

            if args.dry_run:
                print(f"Would process {len(apps)} apps:")
                for a in apps:
                    print(f"  {a['name']} (id={a.get('legacy_id')})")
                return

            all_sr: list[ScanResult] = []; all_mr: list[ModuleRow] = []
            all_fr: list[FileRow] = []; all_rr: list[RecommendationRow] = []
            all_ai: list[AggIssue] = []; all_tr: list[TrendRow] = []
            lock = threading.Lock()
            delay_lock = threading.Lock()

            def _do_app(idx_app: tuple[int, dict]) -> None:
                idx, app = idx_app
                log.info("[%d/%d] %s (id=%s)", idx, len(apps), app["name"], app.get("legacy_id"))
                try:
                    sr, mr, fr, rr, ai, tr = _process_app(
                        client, app, args.skip_no_scan, args.include_sandboxes,
                        skip_checks, prev_data, resume_keys)
                    with lock:
                        all_sr.extend(sr); all_mr.extend(mr); all_fr.extend(fr)
                        all_rr.extend(rr); all_ai.extend(ai); all_tr.extend(tr)
                    for s in sr:
                        sb = f' [{s.sandbox}]' if s.sandbox else ""
                        log.info("    %s | Issues: %d%s", s.health, s.total_issues, sb)
                except AuthError as e:
                    log.error("Authentication failed: %s", e)
                    raise
                except Exception as e:
                    log.warning("    [!] Failed: %s", e)
                    with lock:
                        all_sr.append(_empty_result(app, app.get("legacy_id",0)))
                if args.delay > 0:
                    with delay_lock:
                        time.sleep(args.delay)

            if args.parallel > 1:
                with ThreadPoolExecutor(max_workers=args.parallel) as pool:
                    futs = {pool.submit(_do_app, (i, a)): a for i, a in enumerate(apps, 1)}
                    for fut in as_completed(futs):
                        try: fut.result()
                        except AuthError: raise
                        except Exception as e: log.warning("Worker error: %s", e)
            else:
                for i, app in enumerate(apps, 1):
                    _do_app((i, app))

            # Build output rows
            h_rows = [s.to_row() for s in all_sr]
            m_rows = [m.to_row() for m in all_mr]
            f_rows = [f.to_row() for f in all_fr]
            r_rows = [r.to_row() for r in all_rr]
            t_rows = [t.to_row() for t in all_tr]
            agg = _build_aggregation(all_ai, len(h_rows))

            log.info("\n[*] Writing %d health / %d module / %d file / %d rec / %d trend rows...",
                     len(h_rows), len(m_rows), len(f_rows), len(r_rows), len(t_rows))

            fmt = args.output_format
            if fmt == "xlsx":
                write_excel(h_rows, m_rows, f_rows, r_rows, t_rows, agg, args.output)
            elif fmt == "csv":
                write_csv(h_rows, m_rows, f_rows, r_rows, t_rows, agg, args.output)
            elif fmt == "json":
                write_json(h_rows, m_rows, f_rows, t_rows, args.output)

            _print_summary(h_rows)

    except AuthError as e:
        log.error("FATAL: %s", e)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
